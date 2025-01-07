
# ------------------------------------------------------------
# Import Packages
# ------------------------------------------------------------

# Import packages after ensuring installation
import googlemaps
import geocoder
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Import standard library packages
import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog
from datetime import datetime, timedelta
import os


# ------------------------------------------------------------
# Google Maps API Key Handling
# ------------------------------------------------------------



def get_google_maps_api_key():
    """
    Get API Key from the user without storing it in a file.
    The program prompts the user for the key every time it is run.
    """
    # Prompt the user to enter the API key
    api_key = simpledialog.askstring("Enter API Key", "Please enter your Google Maps API key:")
    if api_key:
        print("API key received.")
    else:
        raise ValueError("API key is required!")

    return api_key


# Get the Google Maps API key from the user (if not saved)
GOOGLE_MAPS_API_KEY = get_google_maps_api_key()

# Initialize the Google Maps client with the user's API key
gmaps = googlemaps.Client(key=GOOGLE_MAPS_API_KEY)

# ------------------------------------------------------------
# Functions for Appointment Assignment
# ------------------------------------------------------------

def zip_to_city(zipcode):
    """
    Takes a ZIP code as input and returns the corresponding city using the Google Maps Geocoding API.

    Parameters:
        zipcode (str): A valid ZIP code to query for city information.

    Returns:
        str: The name of the city corresponding to the ZIP code, or "City not found" if the request fails.
    
    """
    g = geocoder.google(f"{zipcode}", key=GOOGLE_MAPS_API_KEY)

    if g.ok:
        return g.city
    else:
        return "City not found"

## Function Assigning Appointments to Closest Sales Rep
def assign_to_closest_rep(appointments, sales_reps, api_key, time_window_start,time_window_end):
    """
    Assigns sales representatives to appointments based on proximity, considering the drive time 
    between the appointment's location and the rep's location. 

    Description:
    The function uses the Google Maps API to calculate the driving distance and duration (in minutes) 
    between the appointment's zip code and each rep's current zip code. The rep with the shortest 
    drive time is assigned to the appointment. The function also ensures no rep is double-booked 
    by checking if they are already assigned to an appointment within the same time window.

    Parameters:
        appointments (list): A list of dictionaries containing details about each appointment, 
                              including 'apptdate' (date), 'Zip' (string), 'productid' (ID of the product), 
                              and 'custnumber' (customer ID).
        sales_reps (dict): A dictionary mapping each sales rep to their details, including 'curr_zip' (current zip code) 
                           and 'scope' (the product scope that the rep handles).
        api_key (str): The Google Maps API key used for drive time calculation.
        time_window_start (datetime): A datetime object representing the start of the time window for appointments.
        time_window_end (datetime): A datetime object representing the end of the time window for appointments.

    Function Operation:
    1. The function first filters the appointments based on the provided time window (`time_window_start` to `time_window_end`).
    2. For each appointment in the time window, it calculates the drive time from the appointment's zip code 
       to each rep's zip code using the Google Maps API.
    3. It assigns the appointment to the rep with the shortest drive time, provided the rep is eligible based 
       on their product scope.
    4. If any rep has more than one appointment assigned in the same time window, the function prevents double booking 
       by selecting the appointment with the shortest drive time and unassigning others.
    5. It then reassigns unassigned appointments to available reps, following the same drive time and eligibility checks.

    Returns:
    A list of dictionaries, each containing the appointment details, the assigned rep, and the corresponding drive time 
    for each assignment. If an appointment could not be assigned, a message is printed, but the process continues.
    """
    # Initialize list to store appointment assignments
    appointment_assignments = []

    
     # Track which reps have been assigned to which appointments
    rep_appointments = {rep: [] for rep in sales_reps}

    # Cache for storing drive times between zip codes
    drive_time_cache = {}  # Global cache

    def get_drive_time(zip1, zip2):
        """Compute drive time using Google Maps API and cache results."""
        # Ensure pairs are stored consistently regardless of order
        cache_key = tuple(sorted([zip1, zip2]))  # Sort the zip codes to create a consistent key

        if cache_key in drive_time_cache:
            return drive_time_cache[cache_key]  # Return cached value if already computed

        # Make API call if not in cache
        directions_result = gmaps.distance_matrix(origins=zip1, destinations=zip2, mode="driving")

        if directions_result["status"] == "OK":
            drive_time = directions_result["rows"][0]["elements"][0]["duration"]["value"] / 60  # Convert to minutes
            drive_time_cache[cache_key] = drive_time  # Save to cache for future use
            return drive_time
        else:
            # Log an error if the API call fails
            print(f"API call failed for zip pair: {zip1} -> {zip2}")
            return None  # Return None if API call fails

    # Convert appointment times to datetime objects if they are in string format
    for appointment in appointments:
        appointment["apptdate"] = datetime.strptime(appointment["apptdate"], "%m/%d/%Y %H:%M:%S") if isinstance(appointment["apptdate"], str) else appointment["apptdate"]

    
    # Filter appointments that fall within the specified time window
    filtered_appointments = [
        appt for appt in appointments
        if time_window_start <= appt["apptdate"] <= time_window_end
    ]

    # Step 1: Initial assignment of appointments by closest available rep
    for appointment in filtered_appointments:
        appt_zip = appointment["Zip"]
        product_id = appointment["productid"]

        closest_rep = None
        min_drive_time = float('inf')  # Start with a very high value to find the minimum

        for rep, details in sales_reps.items():
            if product_id in details["scope"]:  # Check if the rep scope is eligible for appointment product
                rep_zip = details["curr_zip"]
                drive_time = get_drive_time(appt_zip, rep_zip)

                if drive_time is not None and drive_time < min_drive_time:
                    min_drive_time = drive_time
                    closest_rep = rep  # Assign the closest rep

        if closest_rep:
            rep_appointments[closest_rep].append(appointment)
            appointment_assignments.append({
                "appointment": appointment,
                "assigned_to": closest_rep,
                "drive_time": min_drive_time
            })
        else:
            print(f"No eligible rep found for appointment {appointment['custnumber']} ({appt_zip})")
            continue

    # Step 2: Adjust assignments to avoid double booking
    unassigned_appointments = []
    for rep, assigned_appts in rep_appointments.items():
        if len(assigned_appts) > 1:
            sorted_appts = sorted(
                [appt for appt in appointment_assignments if appt["assigned_to"] == rep],
                key=lambda x: x["drive_time"]
            )

            kept_appointment = sorted_appts.pop(0)  # Keep the appointment with the shortest drive time
            for extra_appt in sorted_appts:
                unassigned_appointments.append(extra_appt["appointment"])
                appointment_assignments.remove(extra_appt)

    # Step 3: Reassign unassigned appointments to available reps
    for appointment in unassigned_appointments:
        appt_zip = appointment["Zip"]
        product_id = appointment["productid"]

        closest_rep = None
        min_drive_time = float('inf')

        for rep, details in sales_reps.items():
            if product_id in details["scope"] and not rep_appointments[rep]:
                rep_zip = details["curr_zip"]
                drive_time = get_drive_time(appt_zip, rep_zip)

                if drive_time is not None and drive_time < min_drive_time:
                    min_drive_time = drive_time
                    closest_rep = rep

        if closest_rep:
            rep_appointments[closest_rep].append(appointment)
            appointment_assignments.append({
                "appointment": appointment,
                "assigned_to": closest_rep,
                "drive_time": min_drive_time
            })
        else:
            print(f"Could not assign appointment {appointment['custnumber']} ({appt_zip}).")

    return appointment_assignments

def update_sales_reps_zip(appointment_assignments, sales_reps):
    """
    Updates the current zip code of sales reps based on their last assigned appointment.

    Parameters:
        appointment_assignments (list): List of dictionaries containing appointment details and assigned reps.
        sales_reps (dict): Dictionary of sales reps and their details.

    Returns:
        dict: The updated `sales_reps` dictionary with rep's current ZIP code (`curr_zip`) set to the ZIP code of their
              most recent appointment.
    """
    # Group assignments by rep
    rep_assignments = {}
    for assignment in appointment_assignments:
        rep = assignment["assigned_to"]
        if rep not in rep_assignments:
            rep_assignments[rep] = []
        rep_assignments[rep].append(assignment["appointment"])

    # Update each rep's current zip based on the last appointment
    for rep, assignments in rep_assignments.items():
        # Sort assignments by appointment time
        sorted_assignments = sorted(assignments, key=lambda x: x["apptdate"])
        # Get the zip code of the last appointment
        last_appt = sorted_assignments[-1]
        sales_reps[rep]["curr_zip"] = last_appt["Zip"]

        
    
    

    return sales_reps

def process_time_window(window_start, window_end, appointments, sales_reps, google_maps_api_key, window_number):
    """
    Assigns appointments to sales representatives (based on proximity) within a given time window.

    This function filters appointments that fall within the specified time window, then assigns each appointment to the 
    sales rep with the shortest drive time using the Google Maps API. The details of each assignment, including the 
    appointment number, city, appointment time, assigned representative, and drive time, are printed for the specified 
    time window.

    Parameters:
        window_start (datetime): The start datetime of the time window for assigning appointments.
        window_end (datetime): The end datetime of the time window for assigning appointments.
        appointments (list): A list of dictionaries, each containing appointment details, including customer number, 
                             appointment time, and ZIP code.
        sales_reps (dict): A dictionary where each key is a sales rep's name or ID, and the value is a dictionary containing 
                           the rep's details, such as current ZIP code and eligibility.
        google_maps_api_key (str): The Google Maps API key used to calculate drive times between appointments and sales reps.
        window_number (int): The number of the current time window being processed (e.g., 1, 2, 3).

    Returns:
        list: A list of appointment assignments, where each entry is a dictionary containing the appointment details, 
              the assigned rep, and drive time.
    """
    assignments = assign_to_closest_rep(appointments, sales_reps, google_maps_api_key, time_window_start=window_start, time_window_end=window_end)

    # Print the time window and the assignments
    print(f"Time Window {window_number}: {window_start} to {window_end}\n")
    
    for assignment in assignments:
        appt = assignment["appointment"]
        rep = assignment["assigned_to"]
        drive_time = assignment.get("drive_time", "Unknown")
        customer_number = appt.get("custnumber")
        appt_date = appt.get("apptdate").strftime("%H:%M")
        zip_code = appt.get("Zip")
        city_name = zip_to_city(zip_code)
    
    print("\n")
    return assignments

# Modify sales reps between second and third time windows
def modify_sales_reps(sales_reps, window_number,time_window_start, time_window_end):
    """
    Displays the current list of sales representatives assigned to a specific time window, and provides options for 
    the user to modify the sales reps by adding, removing, or retaining them. The function continues prompting the user 
    until a valid action is selected, and it returns the updated sales reps dictionary.

    Parameters:
        sales_reps (dict): A dictionary of current sales representatives where the keys are representative names or IDs 
                           and the values are their associated details (e.g., zip code, scope).
        window_number (int): The number of the current time window (e.g., 1, 2, 3) for which the reps are being modified.
        time_window_start (datetime): The start time of the current time window.
        time_window_end (datetime): The end time of the current time window.

    Returns:
        dict: The updated dictionary of sales representatives after modifications.

    Process:
        - Displays the list of current sales reps for the given time window.
        - Prompts the user to add new sales reps, remove existing ones, or keep the current list.

    """
    while True:
        # Format the time to display only hours and minutes
        start_time = time_window_start.strftime("%I:%M %p")
        end_time = time_window_end.strftime("%I:%M %p")
        
        # Prepare the current sales reps for display
        reps_info = f"---Window {window_number} Assigned---\n ({start_time} - {end_time}) \n\n Current Sale Representatives: \n"
        # Display sales representatives assigned to the current time window
        if not sales_reps:
            reps_info += "No sales reps available."
        else:
            for name, details in sales_reps.items():
                reps_info += f"{name}\n"


        # Show the current sales reps and ask for user action
        action = messagebox.askquestion(
            "Modify Sales Reps",
            f"{reps_info}\n\nDo you want to modify the sales reps for Window {window_number + 1}?",
            icon="question"
        )

        if action == "no":
            return sales_reps  # Exit without modifying

        # Prompt the user for action (add/remove/keep)
        user_action = simpledialog.askstring(
            "Modify Sales Reps",
            "What would you like to do? (add/remove):"
        )

        if not user_action:
            continue

        user_action = user_action.strip().lower()
        # Ask the user for action: add or remove sales reps
        if user_action == "add":
            new_reps_input = simpledialog.askstring(
                "Add Sales Reps",
                "Enter new sales reps in format 'Name, Zip, Scope1;Scope2':"
            )
            if new_reps_input:
                try:
                    new_reps = parse_input_to_sales_reps(new_reps_input) # Parse new reps

                    for name, details in new_reps.items():
                        if name not in sales_reps:
                            messagebox.showinfo("Added", f"Added new sales rep: {name}")
                        else:
                            messagebox.showwarning("Exists", f"Sales rep '{name}' already exists")
                        
                        sales_reps[name] = details

                except Exception as e:
                    messagebox.showerror("Error", f"Error adding sales reps: {e}")

        elif user_action == "remove":
            if not sales_reps:
                messagebox.showwarning("No Reps", "No sales reps to remove.")
                continue
            remove_reps_input = simpledialog.askstring(
                "Remove Sales Reps",
                "Enter sales reps' names to remove (comma-separated):"
            )
            if remove_reps_input:
                remove_reps = [name.strip() for name in remove_reps_input.split(",")]
                for rep in remove_reps:
                    if rep in sales_reps:
                        del sales_reps[rep]
                        messagebox.showinfo("Removed", f"Removed sales rep: {rep}")
                    else:
                        messagebox.showwarning("Not Found", f"Sales rep '{rep}' not found.")

        else:
            messagebox.showerror("Invalid Input", "Please enter 'add', 'remove', or 'quit'.")

    return sales_reps


def run_workflow(appointments, sales_reps, time_window_start, time_window_end, window_number, allow_modify_reps=False):
    """
    Processes appointments for a specific time window, assigns them to sales reps,updates the sales reps with the latest assignment details,
    and allows modifications to the sales reps for the next window.

    Parameters:
        appointments (list): A list of dictionaries, each containing details of an appointment (custnumber, apptdate, Zip, productid, dsp_id).
        sales_reps (dict): A dictionary where each key is a sales rep's name or ID, and each value contains their details 
                           (current zip code, scope).
        time_window_start (datetime): The start of the time window during which appointments are processed.
        time_window_end (datetime): The end of the time window during which appointments are processed.
        window_number (int): A unique identifier for the time window being processed (e.g., 1 for the first time window).
        allow_modify_reps (bool): A flag to determine whether the sales reps can be modified after processing this window. 
                                   Defaults to False (no modification).

    Returns:
        tuple: A tuple containing:
            - sales_reps (dict): The updated dictionary of sales reps after updating current zip (if necessary.
            - results (list): A list of dictionaries containing details of the appointments and their assigned sales reps 

    Process:
        1. Prints details of the current time window being processed.
        2. Assigns appointments to the closest sales reps using the `process_time_window` function.
        3. Updates the sales reps' current zip code based on their last assigned appointment via `update_sales_reps_zip`.
        4. If `allow_modify_reps` is set to True - modifies the list of sales reps after the window.
        5. Collects appointment assignment results and prints details to the console.
        6. Returns the updated sales reps dictionary and list of appointment assignment results.

    """
    # Initialize an empty list to store all the appointment assignment messages
    all_messages = []

    print(f"\nProcessing Time Window {window_number} ({time_window_start} to {time_window_end})...\n")
    
    # Process and assign appointments for the time window
    assignments = process_time_window(time_window_start, time_window_end, appointments, sales_reps, GOOGLE_MAPS_API_KEY, window_number)
    
    # Update sales reps with the last assigned appointment zip code
    sales_reps = update_sales_reps_zip(assignments, sales_reps)


    # Allow modification of sales reps after the first window if enabled
    if allow_modify_reps:
        sales_reps = modify_sales_reps(sales_reps, window_number,time_window_start, time_window_end)

    # Collect results for export
    results = []
    for assignment in assignments:
        appt = assignment["appointment"]
        rep = assignment["assigned_to"]
        drive_time = assignment.get("drive_time", "Unknown")
        customer_number = appt.get("custnumber")
        appt_date = appt.get("apptdate")
        zip_code = appt.get("Zip")
        product_id = appt.get("productid")
        city_name = zip_to_city(zip_code)
        
        # Format the print message and append to results
        message = f"Appointment #{customer_number} ({city_name}) at {appt_date.strftime("%m/%d/%Y")} {appt_date.strftime("%I:%M %p")} assigned to {rep}, Drive Time: {drive_time if drive_time == 'Unknown' else drive_time:.2f} minutes"
        print(message)

        results.append({
            "Customer Number": customer_number,
            "Date": appt.get("apptdate").strftime("%m/%d/%Y"),
            "Time": appt.get("apptdate").strftime("%I:%M %p"),
            "City": city_name,
            "Product": product_id, 
            "Assigned Rep": rep,
            "Drive Time (minutes)": drive_time if drive_time == "Unknown" else f"{drive_time:.2f}"
        })



    print("\n")
    return sales_reps, results




def main_workflow(appointments, sales_reps):
    """
    Orchestrates the processing of appointments across multiple time windows, 
    assigns them to sales reps, and generates an Excel report of the results.

    Parameters:
        appointments (list): A list of dictionaries, each representing an appointment with keys such as:
                             - "apptdate": The appointment date and time (datetime or string in "%m/%d/%Y %H:%M:%S" format).
                             - Other keys relevant to the appointment (e.g., customer info, zip code, etc.).
        sales_reps (dict): A dictionary of sales representatives, with keys as rep names or IDs and values containing 
                           their details (e.g., current zip code, scope).

    Process:
        1. **Prepare Data**: Ensures appointment dates are in `datetime` format for consistent processing.
        2. **Define Time Windows**: Splits the appointment schedule into multiple time windows, each with a start and 
           end time, along with a flag to allow modifying sales reps after the window.
        3. **Process Each Time Window**:
            - Calls `run_workflow` for each time window to assign appointments and optionally modify sales reps.
            - Aggregates results for each window into a final report structure.
        4. **Export Results to Excel**:
            - Converts the aggregated results into a DataFrame.
            - Exports the data to an Excel file (`appointments_results.xlsx`) with proper column formatting.
            - Applies bold formatting to rows marking time window titles for clarity.
        5. **Notify Completion**: Prints a message indicating successful export.

    Returns:
        None. Results are exported to an Excel file named `appointments_results.xlsx`.

    Example:
        appointments = [
            {"apptdate": "12/20/2024 09:00:00", "custnumber": 101, "Zip": "98001", "productid": "product1"},
            {"apptdate": "12/20/2024 11:00:00", "custnumber": 102, "Zip": "98002", "productid": "product2"},
        ]
        sales_reps = {
            "Rep A": {"zip": "98001", "scope": ["product1", "product2"]},
            "Rep B": {"zip": "98002", "scope": ["product1", "product3"]},
        }

        main_workflow(appointments, sales_reps)
    """
    try:
        all_results = []

        # Prepare appointment dates if not already formatted
        for appointment in appointments:
            if isinstance(appointment["apptdate"], str):
                appointment["apptdate"] = datetime.strptime(appointment["apptdate"], "%m/%d/%Y %H:%M:%S")

        # Determine the earliest and latest appointment times
        earliest_time = min(appointment["apptdate"] for appointment in appointments)
        last_time = max(appointment["apptdate"] for appointment in appointments)

        # Define time windows
        time_windows = [
            (1, earliest_time, earliest_time + timedelta(hours=2), True),
            (2, earliest_time + timedelta(hours=2, minutes=1), earliest_time + timedelta(hours=5), True),
            (3, earliest_time + timedelta(hours=5, minutes=1), last_time, False),
        ]

        # Process each time window
        for window_number, start, end, allow_modify_reps in time_windows:
            sales_reps, results = run_workflow(
                appointments, sales_reps, start, end, window_number, allow_modify_reps
            )

            # Add time window title and results to the final list
            all_results.append({
                "Customer Number": f"Window {window_number} -- ",
                "Date": "", "Time": "", "City": "", "Product": "", "Assigned Rep": "", "Drive Time (minutes)": ""
            })
            all_results.extend(results)
            all_results.append({
                "Customer Number": "", "Date": "", "Time": "", "City": "", "Product": "", "Assigned Rep": "", "Drive Time (minutes)": ""
            })

        # Convert results to DataFrame
        df_results = pd.DataFrame(all_results)

        # Reorder columns
        df_results = df_results[["Customer Number", "Date", "Time", "City", "Product", "Assigned Rep", "Drive Time (minutes)"]]

        formatted_date = earliest_time.strftime("%Y-%m-%d")

        # Define a writable directory
        writable_dir = os.path.expanduser("~/Documents")  # User's Documents folder
        if not os.path.exists(writable_dir):
            writable_dir = os.getcwd()  # Fall back to current working directory

        # Export to Excel
        file_path = os.path.join(writable_dir, f"appointments_results_{formatted_date}.xlsx")
        try:
            df_results.to_excel(file_path, index=False, engine='openpyxl')
        except Exception as e:
            raise PermissionError(f"Unable to write the file to {file_path}. Check permissions. Error: {e}")

        # Apply bold formatting to time window titles
        try:
            wb = load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    if cell.value and "Window" in str(cell.value):
                        cell.font = Font(bold=True)
            wb.save(file_path)
        except Exception as e:
            raise IOError(f"Failed to modify the Excel file formatting. Error: {e}")

        print(f"Results successfully exported to {file_path}")

    except FileNotFoundError as e:
        print(f"Error: File or directory not found. Ensure the specified path exists. Details: {e}")

    except PermissionError as e:
        print(f"Error: Permission denied. Cannot write to the specified directory or file. Details: {e}")

    except ValueError as e:
        print(f"Error: Invalid data format encountered. Details: {e}")

    except Exception as e:
        print(f"An unexpected error occurred: {e}")






# ------------------------------------------------------------
# Functions for User Input
# ------------------------------------------------------------


def parse_input_to_sales_reps(input_text):
    """
    Parses user-inputted text to create a dictionary of sales representatives with their details.
    
    Parameters:
        input_text (str): Multi-line string where each line represents a sales representative 
                          in the format: 'Name, Zip Code, Scope1; Scope2; Scope 3 Name, Zip Code, Scope1; Scope2; ...'
                          - Name: Representative's name.
                          - Zip Code: The representative's associated zip code.
                          - Scope: Semicolon-separated list of products or services they handle.
                          
                          Multiple representatives can be included on the same line, separated by spaces.
    
    Returns:
        dict: A dictionary containing sales representative details. 
              Example:
              {
                  "Rep A": {"curr_zip": "98001", "scope": ["product1", "product2", "product3"]},
                  "Rep B": {"curr_zip": "98002", "scope": ["product1" "product2",]}
              }
    """
    sales_reps = {}
    lines = input_text.splitlines()

    for line in lines:
        try:
            # Split the line by commas and remove extra spaces
            data = [item.strip() for item in line.split(',')]
            
            if len(data) < 3:  # Ensure the line has at least 3 parts (Rep Name, Zip, Scope)
                raise ValueError(f"Invalid line format: {line}")
            
            name = data[0]  # Rep name
            zip_code = data[1]  # Zip code
            scope = data[2].strip().split(';')  # Scope items separated by semicolon, strip extra spaces

            # Clean extra spaces in the scope list
            scope = [item.strip() for item in scope]

            # Create the sales rep entry with scope as a list
            sales_reps[name] = {
                "curr_zip": zip_code,
                "scope": scope 
            }
        except Exception as e:
            raise ValueError(f" \n Invalid sales representative input. Please check the format: 'Name, Zip Code, Scope1; Scope2; ...' \n \n {e}")
    return sales_reps


def load_appointments_from_excel():
    """
    Loads appointment data from an Excel file, validates, formats, and processes it.

    Description:
        This function prompts the user to select an Excel file containing appointment data. 
        It reads the data, validates required fields, formats appointment details, and 
        stores them in a global variable `appointments`.

    Workflow:
        1. Displays a file dialog to select an Excel file.
        2. Reads the selected file into a pandas DataFrame.
        3. Validates that required columns exist: 'custnumber', 'apptdate', 'Zip', 'productid', 'dsp_id'.
        4. Ensures all required data is present in each row.
        5. Parses and formats the 'apptdate' column to ensure consistency (MM/DD/YYYY HH:MM:SS format).
        6. Converts valid rows into a list of dictionaries representing individual appointments.
        7. Updates a UI text field with a summary of loaded appointments.

    Parameters:
        None

    Returns:
        list: A list of dictionaries containing formatted appointment data, e.g.,
              [
                  {"custnumber": 123, "apptdate": "12/20/2024 09:00:00", 
                   "Zip": "98001", "productid": "prod123", "dsp_id": "dspA"},
                  ...
              ]
        None: If the loading process is canceled or fails.

    """
    global appointments  # Use the global variable to store appointments
    appointments = []

    # Open file dialog to select an Excel file
    file_path = filedialog.askopenfilename(
        title="Select Appointments Excel File",
        filetypes=(("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*"))
    )
    # Check if the user canceled the file selection
    if not file_path:
        return  # # Exit if no file is selected

    try:
        # Read Excel file into a pandas DataFrame
        df = pd.read_excel(file_path)

        # Validate that required columns are present in the file
        required_columns = ["custnumber", "apptdate", "Zip", "productid", "dsp_id"]
        if not all(col in df.columns for col in required_columns):
            raise ValueError(f"Excel file must contain the following columns: {', '.join(required_columns)}")

        # Convert DataFrame to list of dictionaries (appointments format)
        appointments = []  # Clear previous appointments before loading new ones
        for _, row in df.iterrows():
            # Check for missing data in required fields
            if pd.isna(row["custnumber"]) or pd.isna(row["apptdate"]) or pd.isna(row["Zip"]) or pd.isna(row["productid"]):
                raise ValueError(f"Missing data in row: {row}")

            # Parse the date (ensure there are no extra spaces)
            appt_date_str = str(row["apptdate"]).strip()  # Strip spaces
            try:
                # Parse the date and time using the given format
                parsed_date = datetime.strptime(appt_date_str, "%m/%d/%Y %I:%M:%S %p")
            except ValueError as e:
                raise ValueError(f"Error parsing appointment date for row {row}: {str(e)}")

            # Format the parsed date to include both the date and time in military format
            formatted_date = parsed_date.strftime("%m/%d/%Y %H:%M:%S")  # MM/DD/YYYY HH:MM:SS

            # Create appointment dictionary and append to the list
            appointment = {
                "custnumber": row["custnumber"],
                "apptdate": formatted_date,  # Use formatted date
                "Zip": str(row["Zip"]).strip(), # Remove whitespace from zip
                "productid": str(row["productid"]).strip(),  # Remove whitespace from productid
                "dsp_id": str(row["dsp_id"]).strip()  # Remove whitespace from dsp_id
            }
            appointments.append(appointment)

        # Update the UI with the loaded appointments summary
        appointments_text.delete("1.0", tk.END)
        appointments_text.insert(tk.END, f"Loaded {len(appointments)} appointments from {file_path}\n")
        
        return appointments  # Return appointments only if successful

    except Exception as e:
        # Display error message if loading fails
        messagebox.showerror("Error", f"Failed to load appointments: {str(e)}")
        return None  # Return None in case of failure



def on_submit():
    """
    Handles the submission of sales representatives and initiates the workflow process.

    Description:
        This function manages the user interaction for submitting sales representatives' data 
        and executing the main workflow. It validates the input data, ensures appointment data 
        has been loaded, and triggers the processing workflow. In case of errors, it provides 
        feedback to the user via message boxes.

    Workflow:
        1. Retrieves the sales representatives' input from a text widget.
        2. Validates the input to ensure it is not empty.
        3. Checks if appointments data has already been loaded globally.
        4. Parses the sales representatives' input using the `parse_input_to_sales_reps` function.
        5. Calls the `main_workflow` function with the loaded appointments and sales reps data.
        6. Displays a success message if the workflow completes successfully.
        7. Shows error message with details.

    """

    global appointments  # Use the global variable to store appointments

    # Retrieve and clean the sales reps input from the text field
    sales_reps_input = sales_reps_text.get("1.0", tk.END).strip()

    # Check if sales reps input is provided
    if not sales_reps_input:
        messagebox.showerror("Input Error", "Sales reps are required.")
        return

    try:
        # Check if appointments have already been loaded
        if not appointments:
            # If appointments are not loaded, ask the user to load the file
            messagebox.showerror("Input Error", "Upload Appointments")
            return

        # Parse sales reps input
        sales_reps = parse_input_to_sales_reps(sales_reps_input)

        # Call the function to process the workflow
        main_workflow(appointments, sales_reps)

        # Display success message
        messagebox.showinfo("Success", "---Window 3 Assigned---\n\n Results successfully exported to\n {file_path}")

    except Exception as e:
        # Show error message if an exception occurs
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


# ------------------------------------------------------------
# GUI Application Setup
# ------------------------------------------------------------

"""
GUI Application Setup for Appointment Assignment Workflow

Description:
    This section of the code sets up the graphical user interface (GUI) for the appointment assignment workflow 
    application. It uses Tkinter to create a simple and user-friendly interface where users can:
    1. Upload an Excel file containing appointment data.
    2. Input sales representative information.
    3. Execute the workflow process by clicking a submit button.

Components:
    - A label and button for uploading appointment data from an Excel file.
    - A text field for displaying loaded appointment data or accepting manual input.
    - A label and text field for inputting sales representatives' information, including their zip codes and scope.
    - A submit button to process the entered data and execute the workflow.

Workflow:
    - Users start by uploading appointment data using the "Upload Excel File" button.
    - The loaded data is displayed in the text field for review.
    - Users then input sales representatives' details into the designated text field.
    - Clicking the "Submit" button triggers the `on_submit` function to validate the data and run the workflow.

Tkinter Components:
    - `root`: The main application window.
    - `appointments_label`: Label describing the appointment input section.
    - `load_excel_button`: Button for uploading Excel files.
    - `appointments_text`: Text field for displaying or inputting appointment data.
    - `sales_reps_label`: Label describing the sales reps input section.
    - `sales_reps_text`: Text field for inputting sales representatives' details.
    - `submit_button`: Button to submit the input data and execute the workflow.

Event Loop:
    The `root.mainloop()` function starts the Tkinter event loop, keeping the application responsive to user interactions.

Example Usage:
    1. Click "Upload Excel File" to load appointment data from an Excel file.
    2. Enter sales representatives' details in the format:
        Name, Zip Code, Scope1; Scope2; ...
    3. Click "Submit" to process the data.

"""


# GUI Application
root = tk.Tk()
root.title("Appointment Assignment Workflow")

# Appointment input section
appointments_label = tk.Label(root, text="Enter Appointments \n(custnumber, apptdate, Zip, productid, dsp_id) ")
appointments_label.pack()

# Button to load appointments from Excel
load_excel_button = tk.Button(root, text="Upload Excel File", command=load_appointments_from_excel)
load_excel_button.pack()

# Text field for manual input or confirmation
appointments_text = tk.Text(root, height=10, width=50)
appointments_text.pack()
appointments_text.insert(tk.END, "Click 'Upload Excel File' button.\n ")

# Sales rep input section
sales_reps_label = tk.Label(root, text="Enter Sales Reps\n(Name, Zip Code, Scope1 ; Scope 2)")
sales_reps_label.pack()
sales_reps_text = tk.Text(root, height=10, width=50)
sales_reps_text.pack()

# Insert example text as placeholder
sales_reps_text.insert(tk.END, "Example Input:\n\nJohn, 98026, OLS; Bath; Combo; DEF; HA; Kitch;Shwr; T2S; Tub; WIT \nJane, 98004, OLS; Bath; Combo; DEF; HA; Kitch; Shwr; T2S; Tub; WIT\n")

# Submit button to process data
submit_button = tk.Button(root, text="Submit", command=on_submit)
submit_button.pack()

# Start Tkinter event loop
root.mainloop()